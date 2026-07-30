import os
import random
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import (Flask, flash, g, redirect, render_template, request,
                    send_file, send_from_directory, session, url_for)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, UnidentifiedImageError
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (Image as RLImage, Paragraph, SimpleDocTemplate,
                                 Spacer, Table, TableStyle)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "checkin.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
ALLOWED_EXT = {"png", "jpg", "jpeg", "webp"}

# Se DATABASE_URL estiver definida (Render/Supabase Postgres), o app usa
# Postgres. Sem ela, cai para SQLite local — útil para testar na sua máquina.
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras

    IntegrityError = psycopg2.IntegrityError
else:
    IntegrityError = sqlite3.IntegrityError

# Senha padrão do painel do professor/admin. TROQUE antes de usar de verdade!
# Pode também ser definida pela variável de ambiente ADMIN_PASSWORD.
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "treino123")

# Limites de tentativas de PIN (proteção simples contra tentativa e erro).
MAX_PIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

# Compressão de fotos: qualquer foto enviada é redimensionada e reencodada
# para não pesar o armazenamento (importante em planos gratuitos).
MAX_PHOTO_DIMENSION = 720
PHOTO_JPEG_QUALITY = 82

# Paleta usada para colorir o avatar de quem ainda não tem foto (cor fixa
# por aluno, calculada a partir do id, para reconhecer o cartão dele rápido).
AVATAR_COLORS = [
    "#3a7bc8", "#c0453f", "#3f9463", "#d4af37",
    "#8b5fbf", "#c97a3d", "#4a8fa8", "#a3475f",
]

# Faixas (Hapkido) usadas como indicador de progresso: sobe conforme o
# total de treinos confirmados ao longo da vida do aluno no app.
BELTS = [
    {"min": 0, "name": "Faixa Branca", "color": "#f2f2ee"},
    {"min": 10, "name": "Faixa Amarela", "color": "#e7c548"},
    {"min": 25, "name": "Faixa Verde", "color": "#3f9463"},
    {"min": 50, "name": "Faixa Azul", "color": "#3a7bc8"},
    {"min": 80, "name": "Faixa Vermelha", "color": "#c0453f"},
    {"min": 120, "name": "Faixa Preta", "color": "#1b1b1b"},
]


def belt_info(total_checkins):
    """Retorna dados da faixa atual do aluno e progresso até a próxima."""
    current = BELTS[0]
    idx = 0
    for i, b in enumerate(BELTS):
        if total_checkins >= b["min"]:
            current = b
            idx = i

    if idx == len(BELTS) - 1:
        span = max(1, total_checkins - current["min"] or 1)
        progressed = span
        next_name, remaining = None, 0
    else:
        nxt = BELTS[idx + 1]
        span = nxt["min"] - current["min"]
        progressed = total_checkins - current["min"]
        next_name, remaining = nxt["name"], nxt["min"] - total_checkins

    pct = 100 if idx == len(BELTS) - 1 else max(0, min(100, round(progressed / span * 100)))
    return {
        "name": current["name"],
        "color": current["color"],
        "pct": pct,
        "next_name": next_name,
        "remaining": remaining,
    }

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# --------------------------------------------------------------------------
# Banco de dados — camada fina que funciona com SQLite (dev local) ou
# Postgres (produção), usando sempre "?" como marcador de parâmetro no
# código e traduzindo para "%s" quando o backend é Postgres.
# --------------------------------------------------------------------------
class DB:
    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=()):
        cur = self._conn.cursor()
        if USE_POSTGRES:
            sql = sql.replace("?", "%s")
        cur.execute(sql, params)
        return cur

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()


def _raw_connect():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        return conn
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_db():
    if "db" not in g:
        g.db = DB(_raw_connect())
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                nickname TEXT,
                photo TEXT,
                pin TEXT,
                pin_attempts INTEGER NOT NULL DEFAULT 0,
                pin_locked_until TEXT,
                pin_is_custom INTEGER NOT NULL DEFAULT 0,
                birth_date TEXT,
                real_belt TEXT,
                phone TEXT,
                guardian_name TEXT,
                guardian_phone TEXT,
                address TEXT,
                email TEXT,
                modality TEXT,
                monthly_fee REAL,
                due_day INTEGER,
                billing_status TEXT NOT NULL DEFAULT 'ativo',
                notes TEXT,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS checkins (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES students (id),
                workout_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                reviewed_at TEXT
            );

            CREATE TABLE IF NOT EXISTS champions (
                id SERIAL PRIMARY KEY,
                year INTEGER NOT NULL UNIQUE,
                student_id INTEGER NOT NULL REFERENCES students (id),
                total_checkins INTEGER NOT NULL,
                closed_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS mensalidades (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES students (id),
                ref_month INTEGER NOT NULL,
                ref_year INTEGER NOT NULL,
                valor REAL NOT NULL,
                due_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pendente',
                payment_date TEXT,
                payment_method TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                UNIQUE (student_id, ref_month, ref_year)
            );
            """
        )
        # Migração seguindo para bancos criados antes destas colunas existirem.
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS pin TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS pin_attempts INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS pin_locked_until TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS pin_is_custom INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS birth_date TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS real_belt TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS phone TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS guardian_name TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS guardian_phone TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS address TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS email TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS modality TEXT")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS monthly_fee REAL")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS due_day INTEGER")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS billing_status TEXT NOT NULL DEFAULT 'ativo'")
        cur.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS notes TEXT")
        cur.close()
        conn.close()
        return

    db = sqlite3.connect(DB_PATH)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            nickname TEXT,
            photo TEXT,
            pin TEXT,
            pin_attempts INTEGER NOT NULL DEFAULT 0,
            pin_locked_until TEXT,
            pin_is_custom INTEGER NOT NULL DEFAULT 0,
            birth_date TEXT,
            real_belt TEXT,
            phone TEXT,
            guardian_name TEXT,
            guardian_phone TEXT,
            address TEXT,
            email TEXT,
            modality TEXT,
            monthly_fee REAL,
            due_day INTEGER,
            billing_status TEXT NOT NULL DEFAULT 'ativo',
            notes TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS checkins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            workout_date TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL,
            reviewed_at TEXT,
            FOREIGN KEY (student_id) REFERENCES students (id)
        );

        CREATE TABLE IF NOT EXISTS champions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL UNIQUE,
            student_id INTEGER NOT NULL,
            total_checkins INTEGER NOT NULL,
            closed_at TEXT NOT NULL,
            FOREIGN KEY (student_id) REFERENCES students (id)
        );

        CREATE TABLE IF NOT EXISTS mensalidades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            ref_month INTEGER NOT NULL,
            ref_year INTEGER NOT NULL,
            valor REAL NOT NULL,
            due_date TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pendente',
            payment_date TEXT,
            payment_method TEXT,
            notes TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (student_id) REFERENCES students (id),
            UNIQUE (student_id, ref_month, ref_year)
        );
        """
    )
    # Migração segura: bancos criados antes destas colunas existirem.
    for stmt in (
        "ALTER TABLE students ADD COLUMN pin TEXT",
        "ALTER TABLE students ADD COLUMN pin_attempts INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE students ADD COLUMN pin_locked_until TEXT",
        "ALTER TABLE students ADD COLUMN pin_is_custom INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE students ADD COLUMN birth_date TEXT",
        "ALTER TABLE students ADD COLUMN real_belt TEXT",
        "ALTER TABLE students ADD COLUMN phone TEXT",
        "ALTER TABLE students ADD COLUMN guardian_name TEXT",
        "ALTER TABLE students ADD COLUMN guardian_phone TEXT",
        "ALTER TABLE students ADD COLUMN address TEXT",
        "ALTER TABLE students ADD COLUMN email TEXT",
        "ALTER TABLE students ADD COLUMN modality TEXT",
        "ALTER TABLE students ADD COLUMN monthly_fee REAL",
        "ALTER TABLE students ADD COLUMN due_day INTEGER",
        "ALTER TABLE students ADD COLUMN billing_status TEXT NOT NULL DEFAULT 'ativo'",
        "ALTER TABLE students ADD COLUMN notes TEXT",
    ):
        try:
            db.execute(stmt)
        except sqlite3.OperationalError:
            pass  # coluna já existe
    db.commit()
    db.close()


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


def save_student_photo(file_storage, name):
    """Recebe o arquivo de foto enviado, redimensiona e comprime antes de
    salvar (sempre como .jpg), para não pesar o armazenamento. Retorna o
    nome do arquivo salvo, ou None se não havia foto ou o arquivo é inválido."""
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        return None

    try:
        img = Image.open(file_storage.stream)
        img = img.convert("RGB")
    except UnidentifiedImageError:
        return None

    img.thumbnail((MAX_PHOTO_DIMENSION, MAX_PHOTO_DIMENSION), Image.LANCZOS)

    safe_base = secure_filename(name).lower() or "aluno"
    filename = f"{safe_base}-{int(datetime.now().timestamp())}.jpg"
    img.save(
        os.path.join(app.config["UPLOAD_FOLDER"], filename),
        "JPEG",
        quality=PHOTO_JPEG_QUALITY,
        optimize=True,
    )
    return filename


def generate_pin():
    return f"{random.randint(0, 9999):04d}"


def is_student_verified(student_id):
    """Alunos sem PIN definido (cadastro antigo) passam direto, para não
    trancar ninguém fora. Quem tem PIN precisa ter verificado nesta sessão."""
    student = get_student(student_id)
    if not student or not student["pin"]:
        return True
    return student_id in session.get("verified_students", [])


def mark_student_verified(student_id):
    verified = session.get("verified_students", [])
    if student_id not in verified:
        verified.append(student_id)
    session["verified_students"] = verified


def pin_lock_remaining_minutes(student):
    """Minutos restantes de bloqueio por tentativas erradas, ou 0 se livre."""
    locked_until = student["pin_locked_until"]
    if not locked_until:
        return 0
    until = datetime.fromisoformat(locked_until)
    remaining = (until - datetime.now()).total_seconds()
    if remaining <= 0:
        return 0
    return max(1, int(remaining // 60) + 1)


def register_failed_pin(student_id, current_attempts):
    db = get_db()
    attempts = (current_attempts or 0) + 1
    if attempts >= MAX_PIN_ATTEMPTS:
        locked_until = (datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
        db.execute(
            "UPDATE students SET pin_attempts = 0, pin_locked_until = ? WHERE id = ?",
            (locked_until, student_id),
        )
    else:
        db.execute("UPDATE students SET pin_attempts = ? WHERE id = ?", (attempts, student_id))
    db.commit()
    return attempts


def reset_pin_attempts(student_id):
    db = get_db()
    db.execute(
        "UPDATE students SET pin_attempts = 0, pin_locked_until = NULL WHERE id = ?",
        (student_id,),
    )
    db.commit()


def today_str():
    return date.today().isoformat()


def calculate_age(birth_date_str):
    if not birth_date_str:
        return None
    try:
        b = date.fromisoformat(birth_date_str)
    except ValueError:
        return None
    today = date.today()
    age = today.year - b.year - ((today.month, today.day) < (b.month, b.day))
    return age


def display_name(student):
    return student["nickname"] or student["name"]


# --------------------------------------------------------------------------
# Financeiro / Mensalidades
# --------------------------------------------------------------------------
PAYMENT_METHODS = [
    ("pix", "Pix"),
    ("dinheiro", "Dinheiro"),
    ("cartao_credito", "Cartão de Crédito"),
    ("cartao_debito", "Cartão de Débito"),
    ("transferencia", "Transferência"),
    ("outro", "Outro"),
]
PAYMENT_METHOD_LABELS = dict(PAYMENT_METHODS)

MENSALIDADE_STATUSES = [
    ("pendente", "Pendente"),
    ("pago", "Pago"),
    ("atrasado", "Atrasado"),
    ("isento", "Isento"),
]
MENSALIDADE_STATUS_LABELS = dict(MENSALIDADE_STATUSES)

BILLING_STATUSES = [
    ("ativo", "Ativo"),
    ("inativo", "Inativo"),
    ("trancado", "Trancado"),
]

MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def month_name_pt(month, year=None):
    label = MESES_PT[month] if 1 <= month <= 12 else str(month)
    return f"{label}/{year}" if year else label


def format_money(value):
    if value is None:
        value = 0
    text = f"{float(value):,.2f}"
    text = text.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {text}"


def days_in_month(month, year):
    if month == 12:
        nxt = date(year + 1, 1, 1)
    else:
        nxt = date(year, month + 1, 1)
    return (nxt - date(year, month, 1)).days


def add_months(month, year, delta=1):
    total = (month - 1) + delta
    new_year = year + total // 12
    new_month = total % 12 + 1
    return new_month, new_year


def days_overdue(due_date_str):
    try:
        due = date.fromisoformat(due_date_str)
    except (ValueError, TypeError):
        return 0
    delta = (date.today() - due).days
    return delta if delta > 0 else 0


def refresh_overdue_mensalidades():
    """Mensalidades pendentes cujo vencimento já passou viram 'atrasado'.
    Chamado no início das rotas do módulo financeiro (sem precisar de cron)."""
    db = get_db()
    db.execute(
        "UPDATE mensalidades SET status = 'atrasado' "
        "WHERE status = 'pendente' AND due_date < ?",
        (today_str(),),
    )
    db.commit()


def generate_next_month_mensalidades():
    """Gera a mensalidade do mês seguinte (relativo a hoje) para todo aluno
    com billing_status='ativo' e mensalidade configurada. Não duplica quem
    já tem mensalidade gerada pra esse mês/ano (UNIQUE student+ref_month+ref_year)."""
    db = get_db()
    today = date.today()
    ref_month, ref_year = add_months(today.month, today.year, 1)

    students = db.execute(
        "SELECT * FROM students WHERE billing_status = 'ativo' AND monthly_fee IS NOT NULL"
    ).fetchall()

    created = 0
    for s in students:
        exists = db.execute(
            "SELECT id FROM mensalidades WHERE student_id = ? AND ref_month = ? AND ref_year = ?",
            (s["id"], ref_month, ref_year),
        ).fetchone()
        if exists:
            continue
        due_day = s["due_day"] or 5
        due_day = min(due_day, days_in_month(ref_month, ref_year))
        due_date = date(ref_year, ref_month, due_day).isoformat()
        try:
            db.execute(
                "INSERT INTO mensalidades (student_id, ref_month, ref_year, valor, due_date, status, created_at) "
                "VALUES (?, ?, ?, ?, ?, 'pendente', ?)",
                (s["id"], ref_month, ref_year, s["monthly_fee"], due_date, datetime.now().isoformat()),
            )
            created += 1
        except IntegrityError:
            db.rollback()
    db.commit()
    return created, ref_month, ref_year


def financial_dashboard_data():
    """Indicadores e dados de gráfico pro dashboard de Mensalidades."""
    db = get_db()
    today = date.today()

    total_ativos = db.execute(
        "SELECT COUNT(*) AS total FROM students WHERE billing_status = 'ativo'"
    ).fetchone()["total"]

    prevista = db.execute(
        "SELECT COALESCE(SUM(valor), 0) AS total FROM mensalidades WHERE ref_month = ? AND ref_year = ?",
        (today.month, today.year),
    ).fetchone()["total"]

    recebida = db.execute(
        "SELECT COALESCE(SUM(valor), 0) AS total FROM mensalidades "
        "WHERE ref_month = ? AND ref_year = ? AND status = 'pago'",
        (today.month, today.year),
    ).fetchone()["total"]

    pendente = db.execute(
        "SELECT COALESCE(SUM(valor), 0) AS total FROM mensalidades WHERE status IN ('pendente', 'atrasado')"
    ).fetchone()["total"]

    inadimplentes = db.execute(
        "SELECT COUNT(DISTINCT student_id) AS total FROM mensalidades WHERE status = 'atrasado'"
    ).fetchone()["total"]

    pct_inadimplencia = (inadimplentes / total_ativos * 100) if total_ativos else 0

    # Últimos 6 meses (mais antigo -> mais recente) pros gráficos
    months_seq = []
    m, y = add_months(today.month, today.year, -5)
    for _ in range(6):
        months_seq.append((m, y))
        m, y = add_months(m, y, 1)

    receita_labels, receita_prevista_serie, receita_recebida_serie = [], [], []
    inadimplencia_serie = []
    for m, y in months_seq:
        receita_labels.append(month_name_pt(m)[:3] + f"/{str(y)[2:]}")
        prev = db.execute(
            "SELECT COALESCE(SUM(valor), 0) AS total FROM mensalidades WHERE ref_month = ? AND ref_year = ?",
            (m, y),
        ).fetchone()["total"]
        rec = db.execute(
            "SELECT COALESCE(SUM(valor), 0) AS total FROM mensalidades "
            "WHERE ref_month = ? AND ref_year = ? AND status = 'pago'",
            (m, y),
        ).fetchone()["total"]
        total_mes = db.execute(
            "SELECT COUNT(*) AS total FROM mensalidades WHERE ref_month = ? AND ref_year = ?",
            (m, y),
        ).fetchone()["total"]
        atrasados_mes = db.execute(
            "SELECT COUNT(*) AS total FROM mensalidades WHERE ref_month = ? AND ref_year = ? AND status = 'atrasado'",
            (m, y),
        ).fetchone()["total"]
        receita_prevista_serie.append(round(float(prev), 2))
        receita_recebida_serie.append(round(float(rec), 2))
        inadimplencia_serie.append(round((atrasados_mes / total_mes * 100), 1) if total_mes else 0)

    formas = db.execute(
        "SELECT payment_method, COALESCE(SUM(valor), 0) AS total FROM mensalidades "
        "WHERE status = 'pago' AND payment_method IS NOT NULL "
        "GROUP BY payment_method"
    ).fetchall()
    formas_labels = [PAYMENT_METHOD_LABELS.get(f["payment_method"], f["payment_method"]) for f in formas]
    formas_valores = [round(float(f["total"]), 2) for f in formas]

    return {
        "total_ativos": total_ativos,
        "prevista": round(float(prevista), 2),
        "recebida": round(float(recebida), 2),
        "pendente": round(float(pendente), 2),
        "inadimplentes": inadimplentes,
        "pct_inadimplencia": round(pct_inadimplencia, 1),
        "receita_labels": receita_labels,
        "receita_prevista_serie": receita_prevista_serie,
        "receita_recebida_serie": receita_recebida_serie,
        "inadimplencia_serie": inadimplencia_serie,
        "formas_labels": formas_labels,
        "formas_valores": formas_valores,
    }


def billing_indicator(student_id):
    """🟢 em dia / 🟡 vence hoje / 🔴 atrasado / cinza sem mensalidade, pra
    mostrar na lista de alunos."""
    db = get_db()
    row = db.execute(
        "SELECT * FROM mensalidades WHERE student_id = ? ORDER BY ref_year DESC, ref_month DESC LIMIT 1",
        (student_id,),
    ).fetchone()
    if not row:
        return None
    if row["status"] == "atrasado":
        return "red"
    if row["status"] == "pendente" and row["due_date"] == today_str():
        return "yellow"
    if row["status"] in ("pago", "isento"):
        return "green"
    return "yellow" if row["status"] == "pendente" else None


SITUACAO_LABELS = {
    "green": "Em dia",
    "yellow": "Pendente",
    "red": "Atrasado",
    None: "Sem mensalidade cadastrada",
}

PERIODO_LABELS = {
    "mes_atual": "Mês atual",
    "3m": "Últimos 3 meses",
    "6m": "Últimos 6 meses",
    "12m": "Últimos 12 meses",
    "todo": "Todo o histórico",
}


def period_months_set(periodo, inicio="", fim=""):
    """Retorna (set_de_(mes,ano) ou None-pra-'sem filtro', label)."""
    today = date.today()
    if periodo == "mes_atual":
        return {(today.month, today.year)}, PERIODO_LABELS["mes_atual"]
    if periodo in ("3m", "6m", "12m"):
        n = {"3m": 3, "6m": 6, "12m": 12}[periodo]
        months = set()
        m, y = today.month, today.year
        for _ in range(n):
            months.add((m, y))
            m, y = add_months(m, y, -1)
        return months, PERIODO_LABELS[periodo]
    if periodo == "custom" and inicio and fim:
        try:
            ini_d = date.fromisoformat(inicio)
            fim_d = date.fromisoformat(fim)
        except ValueError:
            return None, PERIODO_LABELS["todo"]
        months = set()
        m, y = ini_d.month, ini_d.year
        guard = 0
        while (y, m) <= (fim_d.year, fim_d.month) and guard < 600:
            months.add((m, y))
            m, y = add_months(m, y, 1)
            guard += 1
        return months, f"{inicio} a {fim}"
    return None, PERIODO_LABELS["todo"]


def filter_mensalidades_by_period(rows, periodo, inicio="", fim=""):
    months, label = period_months_set(periodo, inicio, fim)
    if months is None:
        return list(rows), label
    return [r for r in rows if (r["ref_month"], r["ref_year"]) in months], label


# --- Geração de PDFs (ReportLab) ------------------------------------------
LOGO_PATH = os.path.join(BASE_DIR, "static", "icons", "icon-192.png")

PDF_DARK = rl_colors.HexColor("#14171f")
PDF_GOLD = rl_colors.HexColor("#a97f1f")
PDF_MUTED = rl_colors.HexColor("#6b7280")
PDF_LINE = rl_colors.HexColor("#e2e2e2")
PDF_GREEN_BG = rl_colors.HexColor("#e3f3e9")
PDF_GREEN_TEXT = rl_colors.HexColor("#2f6b49")
PDF_YELLOW_BG = rl_colors.HexColor("#fff6d9")
PDF_YELLOW_TEXT = rl_colors.HexColor("#8a6d00")
PDF_RED_BG = rl_colors.HexColor("#fce4e4")
PDF_RED_TEXT = rl_colors.HexColor("#a23b3b")
PDF_GRAY_BG = rl_colors.HexColor("#eeeeee")


def _pdf_status_colors(status):
    return {
        "pago": (PDF_GREEN_BG, PDF_GREEN_TEXT),
        "pendente": (PDF_YELLOW_BG, PDF_YELLOW_TEXT),
        "atrasado": (PDF_RED_BG, PDF_RED_TEXT),
        "isento": (PDF_GRAY_BG, PDF_MUTED),
    }.get(status, (rl_colors.white, PDF_DARK))


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="AcademyTitle", fontName="Helvetica-Bold", fontSize=17, textColor=PDF_DARK))
    styles.add(ParagraphStyle(name="ReportKicker", fontName="Helvetica-Bold", fontSize=9, textColor=PDF_GOLD, spaceAfter=2))
    styles.add(ParagraphStyle(name="ReportMeta", fontName="Helvetica", fontSize=9, textColor=PDF_MUTED, alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name="SectionHeading", fontName="Helvetica-Bold", fontSize=12, textColor=PDF_DARK, spaceBefore=16, spaceAfter=8))
    styles.add(ParagraphStyle(name="StudentLine", fontName="Helvetica", fontSize=10.5, textColor=PDF_DARK, leading=15))
    styles.add(ParagraphStyle(name="FootNote", fontName="Helvetica-Oblique", fontSize=8, textColor=PDF_MUTED, leading=11, spaceBefore=4))
    return styles


def _pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(PDF_LINE)
    canvas.line(20 * mm, 15 * mm, A4[0] - 20 * mm, 15 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(PDF_MUTED)
    canvas.drawString(20 * mm, 10 * mm, "Elite Hapkido — gerado automaticamente pelo app de check-in e mensalidades")
    canvas.drawRightString(A4[0] - 20 * mm, 10 * mm, f"Página {doc.page}")
    canvas.restoreState()


def _pdf_header_table(styles, subtitle, meta_lines):
    logo_cell = RLImage(LOGO_PATH, width=15 * mm, height=15 * mm) if os.path.exists(LOGO_PATH) else ""
    title_cell = [
        Paragraph("ELITE HAPKIDO", styles["AcademyTitle"]),
        Paragraph(subtitle, styles["ReportKicker"]),
    ]
    meta_cell = [Paragraph(line, styles["ReportMeta"]) for line in meta_lines]
    header = Table(
        [[logo_cell, title_cell, meta_cell]],
        colWidths=[20 * mm, 95 * mm, 55 * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    return header


def build_student_pdf(student, historico, periodo_label):
    buffer = BytesIO()
    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=18 * mm, bottomMargin=20 * mm, leftMargin=20 * mm, rightMargin=20 * mm,
    )
    elements = []

    hoje = date.today().strftime("%d/%m/%Y")
    situacao = SITUACAO_LABELS.get(billing_indicator(student["id"]))
    elements.append(_pdf_header_table(
        styles, "Relatório Financeiro Individual",
        [f"Emitido em {hoje}", f"Período: {periodo_label}"],
    ))

    nome = display_name(student)
    elements.append(Paragraph(
        f"<b>Aluno:</b> {student['name']}" + (f" ({student['nickname']})" if student['nickname'] else ""),
        styles["StudentLine"],
    ))
    elements.append(Paragraph(f"<b>Modalidade:</b> {student['modality'] or '—'}", styles["StudentLine"]))
    elements.append(Paragraph(f"<b>Faixa oficial:</b> {student['real_belt'] or '—'}", styles["StudentLine"]))

    pago = sum(m["valor"] for m in historico if m["status"] == "pago")
    aberto = sum(m["valor"] for m in historico if m["status"] in ("pendente", "atrasado"))
    qtd_atraso = sum(1 for m in historico if m["status"] == "atrasado")

    elements.append(Paragraph("Resumo financeiro", styles["SectionHeading"]))
    resumo_data = [
        ["Valor da mensalidade", format_money(student["monthly_fee"]) if student["monthly_fee"] else "—"],
        ["Dia de vencimento", str(student["due_day"]) if student["due_day"] else "—"],
        ["Situação atual", situacao],
        [f"Total pago no período ({periodo_label})", format_money(pago)],
        ["Total em aberto (pendente + atrasado)", format_money(aberto)],
        ["Mensalidades em atraso", str(qtd_atraso)],
    ]
    resumo = Table(resumo_data, colWidths=[95 * mm, 75 * mm])
    resumo.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 0), (0, -1), PDF_MUTED),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, PDF_LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(resumo)
    elements.append(Paragraph(
        "* \"Situação atual\" reflete a mensalidade mais recente do aluno em qualquer mês (a situação real dele agora). "
        "Os valores e a tabela abaixo mostram só o período selecionado no filtro, então os dois podem não bater — "
        "por exemplo, se a mensalidade em aberto for de um mês fora do período escolhido.",
        styles["FootNote"],
    ))

    elements.append(Paragraph("Histórico de mensalidades", styles["SectionHeading"]))
    header_row = ["Mês/Ano", "Vencimento", "Valor", "Status", "Pagamento", "Forma"]
    table_data = [header_row]
    row_colors = [None]
    for m in historico:
        table_data.append([
            month_name_pt(m["ref_month"], m["ref_year"]),
            m["due_date"],
            format_money(m["valor"]),
            MENSALIDADE_STATUS_LABELS.get(m["status"], m["status"]),
            m["payment_date"] or "—",
            PAYMENT_METHOD_LABELS.get(m["payment_method"], "—"),
        ])
        row_colors.append(_pdf_status_colors(m["status"]))

    if len(table_data) == 1:
        table_data.append(["Nenhuma mensalidade no período selecionado.", "", "", "", "", ""])
        row_colors.append((rl_colors.white, PDF_MUTED))

    hist_table = Table(table_data, colWidths=[26 * mm, 26 * mm, 24 * mm, 26 * mm, 26 * mm, 26 * mm], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, PDF_LINE),
    ]
    for i, rc in enumerate(row_colors):
        if i == 0 or rc is None:
            continue
        bg, txt = rc
        style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
        style_cmds.append(("TEXTCOLOR", (3, i), (3, i), txt))
    hist_table.setStyle(TableStyle(style_cmds))
    elements.append(hist_table)

    elements.append(Spacer(1, 14))
    elements.append(Paragraph("Totais do período", styles["SectionHeading"]))
    pendente_only = sum(m["valor"] for m in historico if m["status"] == "pendente")
    atrasado_only = sum(m["valor"] for m in historico if m["status"] == "atrasado")
    totais = Table(
        [["Total pago", "Total pendente", "Total em atraso"],
         [format_money(pago), format_money(pendente_only), format_money(atrasado_only)]],
        colWidths=[56.6 * mm, 56.6 * mm, 56.6 * mm],
    )
    totais.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_MUTED),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("TEXTCOLOR", (0, 1), (0, 1), PDF_GREEN_TEXT),
        ("TEXTCOLOR", (1, 1), (1, 1), PDF_YELLOW_TEXT),
        ("TEXTCOLOR", (2, 1), (2, 1), PDF_RED_TEXT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.6, PDF_LINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, PDF_LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(totais)

    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    return buffer


def build_general_report_pdf(rows, periodo_label, total_pago, total_pendente, total_atrasado):
    buffer = BytesIO()
    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=18 * mm, bottomMargin=20 * mm, leftMargin=20 * mm, rightMargin=20 * mm,
    )
    elements = []
    hoje = date.today().strftime("%d/%m/%Y")
    elements.append(_pdf_header_table(
        styles, "Relatório Geral da Academia",
        [f"Emitido em {hoje}", f"Período: {periodo_label}", f"{len(rows)} aluno(s)"],
    ))

    header_row = ["Aluno", "Situação", "Pago", "Pendente", "Atrasado"]
    table_data = [header_row]
    row_colors = [None]
    for r in rows:
        table_data.append([
            r["name"],
            SITUACAO_LABELS.get(r["situacao"]),
            format_money(r["pago"]),
            format_money(r["pendente"]),
            format_money(r["atrasado"]),
        ])
        situ_color = {"green": PDF_GREEN_BG, "yellow": PDF_YELLOW_BG, "red": PDF_RED_BG}.get(r["situacao"], rl_colors.white)
        row_colors.append(situ_color)

    if len(table_data) == 1:
        table_data.append(["Nenhum aluno cadastrado.", "", "", "", ""])
        row_colors.append(rl_colors.white)

    table = Table(table_data, colWidths=[55 * mm, 32 * mm, 30 * mm, 30 * mm, 30 * mm], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, PDF_LINE),
    ]
    for i, bg in enumerate(row_colors):
        if i == 0 or bg is None:
            continue
        style_cmds.append(("BACKGROUND", (1, i), (1, i), bg))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    elements.append(Spacer(1, 14))
    elements.append(Paragraph("Totais consolidados", styles["SectionHeading"]))
    totais = Table(
        [["Total pago", "Total pendente", "Total em atraso"],
         [format_money(total_pago), format_money(total_pendente), format_money(total_atrasado)]],
        colWidths=[56.6 * mm, 56.6 * mm, 56.6 * mm],
    )
    totais.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_MUTED),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("TEXTCOLOR", (0, 1), (0, 1), PDF_GREEN_TEXT),
        ("TEXTCOLOR", (1, 1), (1, 1), PDF_YELLOW_TEXT),
        ("TEXTCOLOR", (2, 1), (2, 1), PDF_RED_TEXT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.6, PDF_LINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, PDF_LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(totais)

    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    return buffer


def get_active_students():
    db = get_db()
    return db.execute(
        "SELECT * FROM students WHERE active = 1 ORDER BY LOWER(name)"
    ).fetchall()


def get_student(student_id):
    db = get_db()
    return db.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()


def counts_for_student(student_id):
    """Retorna (semana, mes, ano, total_vitalicio) de check-ins APROVADOS."""
    db = get_db()
    rows = db.execute(
        "SELECT workout_date FROM checkins WHERE student_id = ? AND status = 'approved'",
        (student_id,),
    ).fetchall()
    today = date.today()
    cur_iso_year, cur_week, _ = today.isocalendar()
    week = month = year = 0
    for r in rows:
        d = date.fromisoformat(r["workout_date"])
        iso_year, iso_week, _ = d.isocalendar()
        if iso_year == cur_iso_year and iso_week == cur_week:
            week += 1
        if d.year == today.year and d.month == today.month:
            month += 1
        if d.year == today.year:
            year += 1
    return week, month, year, len(rows)


def build_ranking(period):
    """period: 'week' | 'month' | 'year'. Retorna lista ordenada de dicts."""
    db = get_db()
    students = get_active_students()
    rows = db.execute(
        "SELECT student_id, workout_date FROM checkins WHERE status = 'approved'"
    ).fetchall()
    today = date.today()
    cur_iso_year, cur_week, _ = today.isocalendar()

    tally = defaultdict(int)
    for r in rows:
        d = date.fromisoformat(r["workout_date"])
        match = False
        if period == "week":
            iso_year, iso_week, _ = d.isocalendar()
            match = iso_year == cur_iso_year and iso_week == cur_week
        elif period == "month":
            match = d.year == today.year and d.month == today.month
        elif period == "year":
            match = d.year == today.year
        if match:
            tally[r["student_id"]] += 1

    ranking = []
    for s in students:
        ranking.append(
            {
                "id": s["id"],
                "name": display_name(s),
                "photo": s["photo"],
                "count": tally.get(s["id"], 0),
            }
        )
    ranking.sort(key=lambda x: (-x["count"], x["name"].lower()))
    return ranking


def is_admin():
    return session.get("is_admin", False)


@app.context_processor
def inject_globals():
    return {
        "is_admin": is_admin(),
        "current_year": date.today().year,
        "AVATAR_COLORS": AVATAR_COLORS,
        "PAYMENT_METHODS": PAYMENT_METHODS,
        "MENSALIDADE_STATUSES": MENSALIDADE_STATUSES,
        "BILLING_STATUSES": BILLING_STATUSES,
        "MESES_PT": MESES_PT,
    }


app.jinja_env.filters["reais"] = format_money


# --------------------------------------------------------------------------
# Rotas públicas / aluno
# --------------------------------------------------------------------------
@app.route("/")
def index():
    students = get_active_students()
    ranking_year = build_ranking("year")[:3]
    return render_template("index.html", students=students, top3=ranking_year)


def attempt_pin(student, pin):
    """Assume que já checamos que a conta não está bloqueada.
    Retorna (sucesso: bool, tentativas_restantes: int|None)."""
    if pin == student["pin"]:
        reset_pin_attempts(student["id"])
        mark_student_verified(student["id"])
        return True, None
    attempts = register_failed_pin(student["id"], student["pin_attempts"])
    left = MAX_PIN_ATTEMPTS - attempts
    return False, left


@app.route("/aluno/<int:student_id>/entrar", methods=["GET", "POST"])
def aluno_entrar(student_id):
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("index"))

    if is_student_verified(student_id):
        return redirect(url_for("aluno_dashboard", student_id=student_id))

    remaining = pin_lock_remaining_minutes(student)
    if remaining:
        flash(
            f"PIN bloqueado após várias tentativas erradas. Tente de novo em "
            f"{remaining} minuto(s), ou peça ajuda ao professor.",
            "error",
        )
        return render_template("aluno_entrar.html", student=student, locked=True)

    if request.method == "POST":
        pin = request.form.get("pin", "").strip()
        ok, left = attempt_pin(student, pin)
        if ok:
            return redirect(url_for("aluno_dashboard", student_id=student_id))
        if left > 0:
            flash(f"PIN incorreto. Mais {left} tentativa(s) antes do bloqueio temporário.", "error")
        else:
            flash(f"PIN incorreto várias vezes. Bloqueado por {LOCKOUT_MINUTES} minutos.", "error")
            return render_template("aluno_entrar.html", student=student, locked=True)

    return render_template("aluno_entrar.html", student=student, locked=False)


@app.route("/entrar", methods=["POST"])
def entrar_por_nome():
    """Login direto pela home: nome (ou apelido) + PIN, sem precisar tocar
    numa foto — usado pelo formulário de entrada no topo da página inicial."""
    nome_digitado = request.form.get("nome", "").strip()
    pin = request.form.get("pin", "").strip()

    if not nome_digitado:
        flash("Digite seu nome para entrar.", "error")
        return redirect(url_for("index"))

    db = get_db()
    student = db.execute(
        "SELECT * FROM students WHERE active = 1 "
        "AND (LOWER(nickname) = LOWER(?) OR LOWER(name) = LOWER(?)) LIMIT 1",
        (nome_digitado, nome_digitado),
    ).fetchone()

    if not student:
        flash("Não encontramos esse nome. Confira a grafia ou peça ajuda ao professor.", "error")
        return redirect(url_for("index"))

    if not student["pin"]:
        # Alunos antigos sem PIN cadastrado continuam entrando direto.
        mark_student_verified(student["id"])
        return redirect(url_for("aluno_dashboard", student_id=student["id"]))

    remaining = pin_lock_remaining_minutes(student)
    if remaining:
        flash(
            f"PIN bloqueado após várias tentativas erradas. Tente de novo em "
            f"{remaining} minuto(s), ou peça ajuda ao professor.",
            "error",
        )
        return redirect(url_for("index"))

    ok, left = attempt_pin(student, pin)
    if ok:
        return redirect(url_for("aluno_dashboard", student_id=student["id"]))
    if left > 0:
        flash(f"PIN incorreto. Mais {left} tentativa(s) antes do bloqueio temporário.", "error")
    else:
        flash(f"PIN incorreto várias vezes. Bloqueado por {LOCKOUT_MINUTES} minutos.", "error")
    return redirect(url_for("index"))


@app.route("/aluno/<int:student_id>")
def aluno_dashboard(student_id):
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("index"))

    if not is_student_verified(student_id):
        return redirect(url_for("aluno_entrar", student_id=student_id))

    week, month, year, total = counts_for_student(student_id)
    belt = belt_info(total)

    db = get_db()
    today_checkin = db.execute(
        "SELECT * FROM checkins WHERE student_id = ? AND workout_date = ?",
        (student_id, today_str()),
    ).fetchone()

    ranking_month = build_ranking("month")
    my_position = next(
        (i + 1 for i, r in enumerate(ranking_month) if r["id"] == student_id), None
    )

    return render_template(
        "aluno_dashboard.html",
        student=student,
        week=week,
        month=month,
        year=year,
        total=total,
        belt=belt,
        today_checkin=today_checkin,
        ranking_month=ranking_month[:5],
        my_position=my_position,
    )


@app.route("/aluno/<int:student_id>/checkin", methods=["POST"])
def fazer_checkin(student_id):
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("index"))

    if not is_student_verified(student_id):
        return redirect(url_for("aluno_entrar", student_id=student_id))

    db = get_db()
    existing = db.execute(
        "SELECT * FROM checkins WHERE student_id = ? AND workout_date = ?",
        (student_id, today_str()),
    ).fetchone()
    if existing:
        flash("Você já registrou o treino de hoje. Aguarde a confirmação do professor.", "info")
    else:
        db.execute(
            "INSERT INTO checkins (student_id, workout_date, status, created_at) VALUES (?, ?, 'pending', ?)",
            (student_id, today_str(), datetime.now().isoformat()),
        )
        db.commit()
        flash("Check-in registrado! Assim que o professor confirmar, ele conta no ranking.", "success")

    return redirect(url_for("aluno_dashboard", student_id=student_id))


@app.route("/aluno/<int:student_id>/pin", methods=["POST"])
def alterar_pin(student_id):
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("index"))

    if not is_student_verified(student_id):
        return redirect(url_for("aluno_entrar", student_id=student_id))

    novo_pin = request.form.get("novo_pin", "").strip()
    confirmar_pin = request.form.get("confirmar_pin", "").strip()

    if not (novo_pin.isdigit() and len(novo_pin) == 4):
        flash("O novo PIN precisa ter exatamente 4 números.", "error")
    elif novo_pin != confirmar_pin:
        flash("Os dois PINs digitados não são iguais. Tente de novo.", "error")
    else:
        db = get_db()
        db.execute(
            "UPDATE students SET pin = ?, pin_is_custom = 1, pin_attempts = 0, pin_locked_until = NULL WHERE id = ?",
            (novo_pin, student_id),
        )
        db.commit()
        flash("PIN atualizado! A partir de agora só você sabe esse número — nem o professor consegue ver.", "success")

    return redirect(url_for("aluno_dashboard", student_id=student_id))


@app.route("/ranking")
def ranking():
    period = request.args.get("periodo", "month")
    if period not in ("week", "month", "year"):
        period = "month"
    data = build_ranking(period)

    db = get_db()
    champion = db.execute(
        "SELECT c.*, s.name, s.nickname, s.photo FROM champions c "
        "JOIN students s ON s.id = c.student_id ORDER BY c.year DESC LIMIT 1"
    ).fetchone()

    return render_template("ranking.html", ranking=data, period=period, champion=champion)


@app.route("/admin/ranking/exportar")
def exportar_ranking():
    if not require_admin():
        return redirect(url_for("admin_login"))

    period = request.args.get("periodo", "month")
    if period not in ("week", "month", "year"):
        period = "month"
    labels = {"week": "Semana", "month": "Mês", "year": "Ano"}
    data = build_ranking(period)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"
    ws.append(["Posição", "Aluno", "Treinos confirmados"])
    header_fill = PatternFill("solid", fgColor="11141B")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(data, start=1):
        ws.append([i, r["name"], r["count"]])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    label = labels.get(period, period)
    filename = f"ranking-elite-hapkido-{label.lower()}-{today_str()}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/sw.js")
def service_worker():
    # Servido na raiz (não em /static/) para que o service worker consiga
    # controlar o app inteiro (scope "/"), permitindo instalar como PWA.
    response = send_from_directory(os.path.join(BASE_DIR, "static"), "sw.js")
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


# --------------------------------------------------------------------------
# Admin (professor)
# --------------------------------------------------------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if senha == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Login realizado.", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Senha incorreta.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("index"))


def require_admin():
    if not is_admin():
        flash("Faça login como professor para acessar essa página.", "error")
        return False
    return True


@app.route("/admin")
def admin_dashboard():
    if not require_admin():
        return redirect(url_for("admin_login"))

    db = get_db()
    pendentes = db.execute(
        "SELECT c.*, s.name, s.nickname, s.photo FROM checkins c "
        "JOIN students s ON s.id = c.student_id "
        "WHERE c.status = 'pending' ORDER BY c.workout_date DESC, c.created_at ASC"
    ).fetchall()

    recentes = db.execute(
        "SELECT c.*, s.name, s.nickname FROM checkins c "
        "JOIN students s ON s.id = c.student_id "
        "WHERE c.status != 'pending' ORDER BY c.reviewed_at DESC LIMIT 15"
    ).fetchall()

    total_alunos = db.execute(
        "SELECT COUNT(*) AS total FROM students WHERE active = 1"
    ).fetchone()["total"]

    return render_template(
        "admin_dashboard.html",
        pendentes=pendentes,
        recentes=recentes,
        total_alunos=total_alunos,
        using_postgres=USE_POSTGRES,
    )


@app.route("/admin/checkin/<int:checkin_id>/aprovar", methods=["POST"])
def aprovar_checkin(checkin_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute(
        "UPDATE checkins SET status = 'approved', reviewed_at = ? WHERE id = ?",
        (datetime.now().isoformat(), checkin_id),
    )
    db.commit()
    flash("Presença confirmada!", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/checkin/<int:checkin_id>/rejeitar", methods=["POST"])
def rejeitar_checkin(checkin_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute(
        "UPDATE checkins SET status = 'rejected', reviewed_at = ? WHERE id = ?",
        (datetime.now().isoformat(), checkin_id),
    )
    db.commit()
    flash("Check-in rejeitado.", "info")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/alunos")
def admin_alunos():
    if not require_admin():
        return redirect(url_for("admin_login"))
    refresh_overdue_mensalidades()
    db = get_db()
    students = db.execute(
        "SELECT * FROM students ORDER BY active DESC, LOWER(name)"
    ).fetchall()
    indicadores = {s["id"]: billing_indicator(s["id"]) for s in students}
    return render_template("admin_alunos.html", students=students, indicadores=indicadores)


@app.route("/admin/alunos/<int:student_id>/editar", methods=["GET", "POST"])
def admin_editar_aluno(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))

    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("admin_alunos"))

    if request.method == "POST":
        birth_date = request.form.get("birth_date", "").strip()
        real_belt = request.form.get("real_belt", "").strip()
        phone = request.form.get("phone", "").strip()
        guardian_name = request.form.get("guardian_name", "").strip()
        guardian_phone = request.form.get("guardian_phone", "").strip()
        address = request.form.get("address", "").strip()
        active = 1 if request.form.get("active") == "on" else 0

        email = request.form.get("email", "").strip()
        modality = request.form.get("modality", "").strip()
        monthly_fee_raw = request.form.get("monthly_fee", "").strip()
        due_day_raw = request.form.get("due_day", "").strip()
        billing_status = request.form.get("billing_status", "ativo").strip()
        notes = request.form.get("notes", "").strip()

        monthly_fee = None
        if monthly_fee_raw:
            try:
                monthly_fee = float(monthly_fee_raw.replace(",", "."))
            except ValueError:
                flash("Valor da mensalidade inválido.", "error")
                return redirect(url_for("admin_editar_aluno", student_id=student_id))

        due_day = None
        if due_day_raw:
            try:
                due_day = int(due_day_raw)
                if not (1 <= due_day <= 28):
                    raise ValueError
            except ValueError:
                flash("Dia de vencimento precisa ser um número entre 1 e 28.", "error")
                return redirect(url_for("admin_editar_aluno", student_id=student_id))

        if birth_date:
            try:
                date.fromisoformat(birth_date)
            except ValueError:
                flash("Data de nascimento inválida.", "error")
                return redirect(url_for("admin_editar_aluno", student_id=student_id))

        db = get_db()

        file = request.files.get("photo")
        if file and file.filename:
            photo_filename = save_student_photo(file, student["name"])
            if photo_filename is None:
                flash("Não consegui processar essa foto (formato inválido). O resto da ficha foi salvo.", "info")
            else:
                db.execute("UPDATE students SET photo = ? WHERE id = ?", (photo_filename, student_id))

        db.execute(
            "UPDATE students SET birth_date = ?, real_belt = ?, phone = ?, "
            "guardian_name = ?, guardian_phone = ?, address = ?, active = ?, "
            "email = ?, modality = ?, monthly_fee = ?, due_day = ?, billing_status = ?, notes = ? "
            "WHERE id = ?",
            (
                birth_date or None,
                real_belt or None,
                phone or None,
                guardian_name or None,
                guardian_phone or None,
                address or None,
                active,
                email or None,
                modality or None,
                monthly_fee,
                due_day,
                billing_status or "ativo",
                notes or None,
                student_id,
            ),
        )
        db.commit()
        flash("Ficha do aluno atualizada!", "success")
        return redirect(url_for("admin_alunos"))

    age = calculate_age(student["birth_date"])
    return render_template("admin_editar_aluno.html", student=student, age=age)


@app.route("/admin/alunos/novo", methods=["POST"])
def admin_novo_aluno():
    if not require_admin():
        return redirect(url_for("admin_login"))

    name = request.form.get("name", "").strip()
    nickname = request.form.get("nickname", "").strip()
    pin = request.form.get("pin", "").strip()
    if not name:
        flash("O nome do aluno é obrigatório.", "error")
        return redirect(url_for("admin_alunos"))

    if pin and (not pin.isdigit() or len(pin) != 4):
        flash("O PIN precisa ter exatamente 4 números. Deixe em branco para gerar automaticamente.", "error")
        return redirect(url_for("admin_alunos"))
    if not pin:
        pin = generate_pin()

    db = get_db()
    db.execute(
        "INSERT INTO students (name, nickname, pin, active, created_at) VALUES (?, ?, ?, 1, ?)",
        (name, nickname or None, pin, datetime.now().isoformat()),
    )
    db.commit()
    flash(f"Aluno {name} cadastrado! PIN de acesso: {pin} (repasse esse número a ele). Adicione foto e mais dados na Ficha completa.", "success")
    return redirect(url_for("admin_alunos"))


@app.route("/admin/alunos/<int:student_id>/novo-pin", methods=["POST"])
def admin_novo_pin(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("admin_alunos"))

    new_pin = generate_pin()
    db = get_db()
    db.execute(
        "UPDATE students SET pin = ?, pin_is_custom = 0, pin_attempts = 0, pin_locked_until = NULL WHERE id = ?",
        (new_pin, student_id),
    )
    db.commit()
    # Invalida a verificação de sessão anterior para esse aluno, se houver.
    verified = session.get("verified_students", [])
    if student_id in verified:
        verified.remove(student_id)
        session["verified_students"] = verified
    flash(f"Novo PIN de {student['nickname'] or student['name']}: {new_pin}", "success")
    return redirect(url_for("admin_alunos"))


@app.route("/admin/alunos/<int:student_id>/inativar", methods=["POST"])
def admin_inativar_aluno(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute("UPDATE students SET active = 0 WHERE id = ?", (student_id,))
    db.commit()
    flash("Aluno removido da lista ativa.", "info")
    return redirect(url_for("admin_alunos"))


@app.route("/admin/alunos/<int:student_id>/reativar", methods=["POST"])
def admin_reativar_aluno(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute("UPDATE students SET active = 1 WHERE id = ?", (student_id,))
    db.commit()
    flash("Aluno reativado.", "success")
    return redirect(url_for("admin_alunos"))


@app.route("/admin/campeoes")
def admin_campeoes():
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    campeoes = db.execute(
        "SELECT c.*, s.name, s.nickname, s.photo FROM champions c "
        "JOIN students s ON s.id = c.student_id ORDER BY c.year DESC"
    ).fetchall()
    ranking_year = build_ranking("year")
    already_closed = any(c["year"] == date.today().year for c in campeoes)
    return render_template(
        "admin_campeoes.html",
        campeoes=campeoes,
        ranking_year=ranking_year,
        already_closed=already_closed,
    )


@app.route("/admin/campeoes/fechar", methods=["POST"])
def fechar_ano():
    if not require_admin():
        return redirect(url_for("admin_login"))

    year = date.today().year
    ranking_year = build_ranking("year")
    if not ranking_year or ranking_year[0]["count"] == 0:
        flash("Ainda não há check-ins aprovados este ano para fechar o ranking.", "error")
        return redirect(url_for("admin_campeoes"))

    winner = ranking_year[0]
    db = get_db()
    try:
        db.execute(
            "INSERT INTO champions (year, student_id, total_checkins, closed_at) VALUES (?, ?, ?, ?)",
            (year, winner["id"], winner["count"], datetime.now().isoformat()),
        )
        db.commit()
        flash(f"Ano {year} fechado! Campeão: {winner['name']} com {winner['count']} treinos.", "success")
    except IntegrityError:
        db.rollback()
        flash(f"O ano {year} já foi fechado anteriormente.", "error")

    return redirect(url_for("admin_campeoes"))


# --------------------------------------------------------------------------
# Financeiro / Mensalidades
# --------------------------------------------------------------------------
@app.route("/admin/mensalidades")
def admin_mensalidades():
    if not require_admin():
        return redirect(url_for("admin_login"))

    refresh_overdue_mensalidades()
    db = get_db()

    aluno_q = request.args.get("aluno", "").strip()
    modalidade_q = request.args.get("modalidade", "").strip()
    status_q = request.args.get("status", "").strip()
    mes_q = request.args.get("mes", "").strip()
    ano_q = request.args.get("ano", "").strip()
    forma_q = request.args.get("forma_pagamento", "").strip()

    today = date.today()
    if not mes_q and not ano_q:
        mes_q, ano_q = str(today.month), str(today.year)

    where = ["1=1"]
    params = []
    if aluno_q:
        where.append("(LOWER(s.name) LIKE ? OR LOWER(s.nickname) LIKE ?)")
        like = f"%{aluno_q.lower()}%"
        params += [like, like]
    if modalidade_q:
        where.append("s.modality = ?")
        params.append(modalidade_q)
    if status_q:
        where.append("m.status = ?")
        params.append(status_q)
    if mes_q:
        where.append("m.ref_month = ?")
        params.append(int(mes_q))
    if ano_q:
        where.append("m.ref_year = ?")
        params.append(int(ano_q))
    if forma_q:
        where.append("m.payment_method = ?")
        params.append(forma_q)

    query = (
        "SELECT m.*, s.name, s.nickname, s.photo, s.modality FROM mensalidades m "
        "JOIN students s ON s.id = m.student_id "
        "WHERE " + " AND ".join(where) +
        " ORDER BY m.due_date ASC"
    )
    mensalidades = db.execute(query, params).fetchall()

    modalidades = db.execute(
        "SELECT DISTINCT modality FROM students WHERE modality IS NOT NULL AND modality != '' ORDER BY modality"
    ).fetchall()

    dashboard = financial_dashboard_data()

    return render_template(
        "admin_mensalidades.html",
        mensalidades=mensalidades,
        dashboard=dashboard,
        modalidades=[m["modality"] for m in modalidades],
        filtros={
            "aluno": aluno_q, "modalidade": modalidade_q, "status": status_q,
            "mes": mes_q, "ano": ano_q, "forma_pagamento": forma_q,
        },
        days_overdue=days_overdue,
        month_name_pt=month_name_pt,
    )


@app.route("/admin/mensalidades/gerar-proximo-mes", methods=["POST"])
def admin_gerar_mensalidades():
    if not require_admin():
        return redirect(url_for("admin_login"))
    created, ref_month, ref_year = generate_next_month_mensalidades()
    if created:
        flash(f"{created} mensalidade(s) de {month_name_pt(ref_month, ref_year)} geradas.", "success")
    else:
        flash(f"Nenhuma mensalidade nova para {month_name_pt(ref_month, ref_year)} (já geradas ou nenhum aluno ativo com valor definido).", "info")
    return redirect(url_for("admin_mensalidades"))


@app.route("/admin/mensalidades/<int:mensalidade_id>/pagar", methods=["POST"])
def admin_pagar_mensalidade(mensalidade_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    payment_date = request.form.get("payment_date", "").strip() or today_str()
    payment_method = request.form.get("payment_method", "").strip() or None
    db = get_db()
    db.execute(
        "UPDATE mensalidades SET status = 'pago', payment_date = ?, payment_method = ? WHERE id = ?",
        (payment_date, payment_method, mensalidade_id),
    )
    db.commit()
    flash("Pagamento registrado!", "success")
    return redirect(request.referrer or url_for("admin_mensalidades"))


@app.route("/admin/mensalidades/<int:mensalidade_id>/isentar", methods=["POST"])
def admin_isentar_mensalidade(mensalidade_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute("UPDATE mensalidades SET status = 'isento' WHERE id = ?", (mensalidade_id,))
    db.commit()
    flash("Mensalidade marcada como isenta.", "success")
    return redirect(request.referrer or url_for("admin_mensalidades"))


@app.route("/admin/mensalidades/<int:mensalidade_id>/editar", methods=["GET", "POST"])
def admin_editar_mensalidade(mensalidade_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    mensalidade = db.execute(
        "SELECT m.*, s.name, s.nickname FROM mensalidades m JOIN students s ON s.id = m.student_id WHERE m.id = ?",
        (mensalidade_id,),
    ).fetchone()
    if not mensalidade:
        flash("Mensalidade não encontrada.", "error")
        return redirect(url_for("admin_mensalidades"))

    if request.method == "POST":
        valor = request.form.get("valor", "").strip()
        due_date = request.form.get("due_date", "").strip()
        status = request.form.get("status", "").strip()
        payment_date = request.form.get("payment_date", "").strip() or None
        payment_method = request.form.get("payment_method", "").strip() or None
        notes = request.form.get("notes", "").strip() or None

        try:
            valor = float(valor.replace(",", "."))
        except ValueError:
            flash("Valor inválido.", "error")
            return redirect(url_for("admin_editar_mensalidade", mensalidade_id=mensalidade_id))

        db.execute(
            "UPDATE mensalidades SET valor = ?, due_date = ?, status = ?, payment_date = ?, "
            "payment_method = ?, notes = ? WHERE id = ?",
            (valor, due_date, status, payment_date, payment_method, notes, mensalidade_id),
        )
        db.commit()
        flash("Mensalidade atualizada!", "success")
        return redirect(url_for("admin_mensalidades"))

    return render_template("admin_editar_mensalidade.html", m=mensalidade, month_name_pt=month_name_pt)


@app.route("/admin/mensalidades/<int:mensalidade_id>/excluir", methods=["POST"])
def admin_excluir_mensalidade(mensalidade_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    db = get_db()
    db.execute("DELETE FROM mensalidades WHERE id = ?", (mensalidade_id,))
    db.commit()
    flash("Registro excluído.", "info")
    return redirect(request.referrer or url_for("admin_mensalidades"))


@app.route("/admin/alunos/<int:student_id>/financeiro")
def admin_perfil_financeiro(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))

    refresh_overdue_mensalidades()
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("admin_alunos"))

    db = get_db()
    historico = db.execute(
        "SELECT * FROM mensalidades WHERE student_id = ? ORDER BY ref_year DESC, ref_month DESC",
        (student_id,),
    ).fetchall()

    week, month, year, total_checkins = counts_for_student(student_id)
    ranking_year = build_ranking("year")
    my_position = next((i + 1 for i, r in enumerate(ranking_year) if r["id"] == student_id), None)

    ultima_presenca = db.execute(
        "SELECT workout_date FROM checkins WHERE student_id = ? AND status = 'approved' "
        "ORDER BY workout_date DESC LIMIT 1",
        (student_id,),
    ).fetchone()

    situacao = billing_indicator(student_id)

    return render_template(
        "admin_perfil_financeiro.html",
        student=student,
        historico=historico,
        age=calculate_age(student["birth_date"]),
        week=week, month=month, year=year, total_checkins=total_checkins,
        my_position=my_position,
        ultima_presenca=ultima_presenca["workout_date"] if ultima_presenca else None,
        situacao=situacao,
        days_overdue=days_overdue,
        month_name_pt=month_name_pt,
        PERIODO_LABELS=PERIODO_LABELS,
    )


@app.route("/admin/alunos/<int:student_id>/pdf")
def admin_aluno_pdf(student_id):
    if not require_admin():
        return redirect(url_for("admin_login"))

    refresh_overdue_mensalidades()
    student = get_student(student_id)
    if not student:
        flash("Aluno não encontrado.", "error")
        return redirect(url_for("admin_alunos"))

    periodo = request.args.get("periodo", "mes_atual")
    inicio = request.args.get("inicio", "")
    fim = request.args.get("fim", "")

    db = get_db()
    historico_all = db.execute(
        "SELECT * FROM mensalidades WHERE student_id = ? ORDER BY ref_year, ref_month",
        (student_id,),
    ).fetchall()
    historico, periodo_label = filter_mensalidades_by_period(historico_all, periodo, inicio, fim)

    try:
        buffer = build_student_pdf(student, historico, periodo_label)
    except Exception:
        app.logger.exception("Falha ao gerar PDF individual (student_id=%s)", student_id)
        flash("Não consegui gerar o PDF agora. Tente novamente em alguns segundos.", "error")
        return redirect(url_for("admin_perfil_financeiro", student_id=student_id))

    filename = f"relatorio-{secure_filename(student['name'])}-{today_str()}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/admin/mensalidades/relatorio-geral")
def admin_relatorio_geral_pdf():
    if not require_admin():
        return redirect(url_for("admin_login"))

    refresh_overdue_mensalidades()
    periodo = request.args.get("periodo", "mes_atual")
    inicio = request.args.get("inicio", "")
    fim = request.args.get("fim", "")

    db = get_db()
    students = db.execute("SELECT * FROM students ORDER BY LOWER(name)").fetchall()
    all_mensalidades = db.execute("SELECT * FROM mensalidades").fetchall()

    by_student = defaultdict(list)
    for m in all_mensalidades:
        by_student[m["student_id"]].append(m)

    rows = []
    total_pago = total_pendente = total_atrasado = 0.0
    label = PERIODO_LABELS.get(periodo, PERIODO_LABELS["todo"])
    for s in students:
        historico = by_student.get(s["id"], [])
        filtrado, label = filter_mensalidades_by_period(historico, periodo, inicio, fim)
        pago = sum(m["valor"] for m in filtrado if m["status"] == "pago")
        pendente = sum(m["valor"] for m in filtrado if m["status"] == "pendente")
        atrasado = sum(m["valor"] for m in filtrado if m["status"] == "atrasado")
        rows.append({
            "name": display_name(s),
            "situacao": billing_indicator(s["id"]),
            "pago": pago, "pendente": pendente, "atrasado": atrasado,
        })
        total_pago += pago
        total_pendente += pendente
        total_atrasado += atrasado

    try:
        buffer = build_general_report_pdf(rows, label, total_pago, total_pendente, total_atrasado)
    except Exception:
        app.logger.exception("Falha ao gerar PDF geral da academia")
        flash("Não consegui gerar o relatório geral agora. Tente novamente em alguns segundos.", "error")
        return redirect(url_for("admin_mensalidades"))

    filename = f"relatorio-geral-elite-hapkido-{today_str()}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


# Cria/atualiza as tabelas do banco assim que o módulo é carregado — precisa
# rodar tanto com "python app.py" (local) quanto com "gunicorn app:app"
# (produção), já que o gunicorn nunca executa o bloco abaixo.
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
